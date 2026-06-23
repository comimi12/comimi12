# -*- coding: utf-8 -*-
"""
월별 매장 분석 자료를 엑셀 한 파일로 내보내기.
집계 + 불만 리뷰 원본 + 개선방안·실행계획 + VOC + CS/QSCS + 추이 그래프.

사용:
  python export_monthly.py            # 현재월
  python export_monthly.py 2026-05
출력: exports/매장리뷰분석_{월}.xlsx
"""
import sys, io, os, json, subprocess, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(HERE, "exports"); os.makedirs(OUTDIR, exist_ok=True)
MONTH = sys.argv[1] if len(sys.argv) > 1 else datetime.date.today().strftime("%Y-%m")

HEAD_FILL = PatternFill("solid", fgColor="14559B")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="0E3C70")
THIN = Side(style="thin", color="D0D6E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def load_js():
    """data.js / ai_notes.js / eck_data.js 를 node로 평가해 JSON 반환."""
    snippet = (
        "global.window={};"
        "const fs=require('fs');"
        "eval(fs.readFileSync('data.js','utf8'));"
        "try{eval(fs.readFileSync('ai_notes.js','utf8'))}catch(e){}"
        "try{eval(fs.readFileSync('eck_data.js','utf8'))}catch(e){}"
        "process.stdout.write(JSON.stringify({reviews:window.REVIEW_DATA||null,ai:window.AI_NOTES||null,eck:window.ECK_DATA||null}));"
    )
    r = subprocess.run(["node", "-e", snippet], cwd=HERE, capture_output=True, text=True, encoding="utf-8")
    if r.returncode != 0:
        raise SystemExit(f"[ERR] JS 로드 실패: {r.stderr[:300]}")
    return json.loads(r.stdout)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def widths(ws, w):
    for i, x in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = x


def main():
    d = load_js()
    R = d.get("reviews") or {}
    AI = d.get("ai") or {}
    ECK = d.get("eck") or {}
    sm = R.get("store_month", {})
    sb = R.get("store_brand", {})
    comps = [c for c in R.get("complaints", []) if c.get("month") == MONTH]

    wb = Workbook()

    # ── 1) 요약 ──
    ws = wb.active; ws.title = "요약"
    ws["A1"] = f"매장 리뷰·VOC 분석 요약 — {MONTH}"; ws["A1"].font = TITLE_FONT
    bm = (R.get("by_month_total", {}) or {}).get(MONTH, {})
    tot = bm.get("total", 0); good = bm.get("칭찬", 0); bad = bm.get("불만", 0)
    rate = round(bad / tot * 100, 1) if tot else 0
    voc = ECK_voc = None
    vrec = [v for v in ((d.get("reviews") or {}).get("voc", {}) or {}).get("records", []) if str(v.get("month")) == MONTH] if isinstance((d.get("reviews") or {}).get("voc"), dict) else []
    # voc 는 reviews.voc 에 있음
    vocobj = R.get("voc") or {}
    vrec = [v for v in vocobj.get("records", []) if v.get("month") == MONTH]
    cs = (ECK.get("cs", {}) or {}).get(MONTH, {})
    qs = (ECK.get("qscs", {}) or {}).get(MONTH, {})
    rows = [
        ("리뷰 총수", tot), ("칭찬", good), ("불만", bad), ("불만율(%)", rate),
        ("VOC 접수", len(vrec)),
        ("CS 점검 대상 매장", cs.get("store_count", "-")),
        ("QSCS 등록/대상", f"{qs.get('reg_count','-')}/{qs.get('roster_count','-')}" if qs else "-"),
    ]
    ws["A3"] = "항목"; ws["B3"] = "값"; style_header(ws, 3, 2)
    for i, (k, v) in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=k).border = BORDER
        ws.cell(row=i, column=2, value=v).border = BORDER
    widths(ws, [22, 18])

    # 월별 추이(브랜드 합) + 라인차트
    months = R.get("months", [])
    bmt = R.get("by_month_total", {})
    ws["D3"] = "월"; ws["E3"] = "리뷰수"; ws["F3"] = "불만"; ws["G3"] = "불만율(%)"
    style_header(ws, 3, 7)
    for i, m in enumerate(months, 4):
        o = bmt.get(m, {}); t = o.get("total", 0); b = o.get("불만", 0)
        ws.cell(row=i, column=4, value=m); ws.cell(row=i, column=5, value=t)
        ws.cell(row=i, column=6, value=b)
        ws.cell(row=i, column=7, value=round(b / t * 100, 1) if t else 0)
    if months:
        ch = LineChart(); ch.title = "월별 리뷰수·불만율 추이"; ch.height = 7; ch.width = 16
        data = Reference(ws, min_col=5, max_col=7, min_row=3, max_row=3 + len(months))
        cats = Reference(ws, min_col=4, min_row=4, max_row=3 + len(months))
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ws.add_chart(ch, "A13")

    # ── 2) 매장별 집계 ──
    ws = wb.create_sheet("매장별")
    ws.append(["브랜드", "매장", "리뷰수", "칭찬", "불만", "불만율(%)"]); style_header(ws, 1, 6)
    rows2 = []
    for store, mm in sm.items():
        o = mm.get(MONTH)
        if not o or not o.get("total"):
            continue
        t = o["total"]; b = o.get("불만", 0)
        rows2.append([sb.get(store, ""), store, t, o.get("칭찬", 0), b, round(b / t * 100, 1) if t else 0])
    rows2.sort(key=lambda r: (-r[5], -r[4]))
    for r in rows2:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        for c in row:
            c.border = BORDER
    widths(ws, [12, 26, 9, 8, 8, 10])
    if rows2:
        ch = BarChart(); ch.title = f"{MONTH} 매장별 불만율 (상위)"; ch.height = 9; ch.width = 20
        n = min(15, len(rows2))
        data = Reference(ws, min_col=6, min_row=1, max_row=1 + n)
        cats = Reference(ws, min_col=2, min_row=2, max_row=1 + n)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        ws.add_chart(ch, "H2")

    # ── 3) 브랜드별 ──
    ws = wb.create_sheet("브랜드별")
    ws.append(["브랜드", "리뷰수", "칭찬", "불만", "불만율(%)"]); style_header(ws, 1, 5)
    bb = {}
    for store, mm in sm.items():
        o = mm.get(MONTH)
        if not o:
            continue
        br = sb.get(store, "")
        a = bb.setdefault(br, [0, 0, 0])
        a[0] += o.get("total", 0); a[1] += o.get("칭찬", 0); a[2] += o.get("불만", 0)
    for br, (t, g, b) in sorted(bb.items(), key=lambda x: -x[1][2]):
        ws.append([br, t, g, b, round(b / t * 100, 1) if t else 0])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        for c in row:
            c.border = BORDER
    widths(ws, [14, 9, 8, 8, 10])

    # ── 4) 불만 리뷰 원본 ──
    ws = wb.create_sheet("불만리뷰")
    ws.append(["브랜드", "매장", "날짜", "불만 내용"]); style_header(ws, 1, 4)
    for c in comps:
        ws.append([c.get("brand", ""), c.get("store", ""), c.get("date", ""), c.get("text", "")])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for c in row:
            c.border = BORDER; c.alignment = WRAP
    widths(ws, [10, 22, 12, 90])

    # ── 5) 개선방안·실행계획 (TOP3) ──
    ws = wb.create_sheet("개선방안_실행계획")
    ws.append(["매장", "불만 요약", "개선 방안", "실행 계획(부서별)"]); style_header(ws, 1, 4)
    for n in (AI.get("monthly", {}) or {}).get(MONTH, []):
        ws.append([
            n.get("store", ""),
            "\n".join(n.get("불만요약", [])),
            "\n".join(n.get("개선방안", [])),
            "\n".join(n.get("완료방법", [])),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for c in row:
            c.border = BORDER; c.alignment = WRAP
    widths(ws, [22, 50, 50, 55])

    # ── 6) VOC ──
    ws = wb.create_sheet("VOC")
    ws.append(["매장", "구분", "문의유형", "접수일", "내용"]); style_header(ws, 1, 5)
    for v in vrec:
        ws.append([v.get("store", ""), v.get("category", ""), v.get("inquiry_type", ""),
                   v.get("reg", ""), v.get("content", "")])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        for c in row:
            c.border = BORDER; c.alignment = WRAP
    widths(ws, [22, 10, 12, 12, 90])

    # ── 7) CS·QSCS ──
    if cs.get("stores"):
        ws = wb.create_sheet("CS체크리스트")
        ws.append(["매장", "완료일수", "점검일수", "점검률(%)"]); style_header(ws, 1, 4)
        days = cs.get("days", [])
        nd = len(days)
        for s in cs["stores"]:
            done = sum(1 for v in (s.get("days", {}) or {}).values() if v)
            ws.append([s.get("store", ""), done, nd, round(done / nd * 100, 1) if nd else 0])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
            for c in row:
                c.border = BORDER
        widths(ws, [26, 10, 10, 10])
    if qs.get("registered") is not None:
        ws = wb.create_sheet("QSCS교육일지")
        ws.append(["매장", "등록여부", "작성일자"]); style_header(ws, 1, 3)
        for r in qs.get("registered", []):
            ws.append([r.get("store", ""), "등록", r.get("date", "")])
        for r in qs.get("not_registered", []):
            ws.append([r.get("store", ""), "미등록", ""])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
            for c in row:
                c.border = BORDER
        widths(ws, [26, 10, 12])

    dest = os.path.join(OUTDIR, f"매장리뷰분석_{MONTH}.xlsx")
    wb.save(dest)
    print(f"[OK] {os.path.relpath(dest, HERE)} 생성 (시트 {len(wb.sheetnames)}개: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
