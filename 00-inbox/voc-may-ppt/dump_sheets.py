# -*- coding: utf-8 -*-
import io
import openpyxl
WD=r"C:\Users\owner\do-better-workspace-v2\00-inbox\voc-may-ppt"
p=r"C:\Users\owner\Desktop\교육팀\0. 온라인리뷰\리뷰집계파일\AI매장리뷰_260611\review_analysis_output\review_analysis_tables_202606111853.xlsx"
wb=openpyxl.load_workbook(p, read_only=True, data_only=True)
o=io.open(WD+r"\sheets_content.txt","w",encoding="utf-8")
# dump small sheets fully, large sheets header+few rows
for ws in wb.worksheets:
    o.write(f"\n{'='*70}\nSHEET: {ws.title}  ({ws.max_row}x{ws.max_column})\n{'='*70}\n")
    maxr = ws.max_row
    limit = maxr if maxr<=95 else 8
    for ri,row in enumerate(ws.iter_rows(values_only=True),1):
        vals=[("" if v is None else str(v)) for v in row]
        # truncate long cells
        vals=[ (v[:60]+"…") if len(v)>60 else v for v in vals]
        o.write(f"R{ri}: {vals}\n")
        if ri>=limit:
            o.write(f"... ({maxr} rows total)\n")
            break
o.close()
print("done")
