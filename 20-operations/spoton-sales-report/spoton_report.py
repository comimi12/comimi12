# -*- coding: utf-8 -*-
"""
SpotOn 일별 매출/객수 리포트 생성기.

사용법:
  python spoton_report.py                      # 어제 하루, 4개 매장 -> Excel
  python spoton_report.py 2026-06-09 2026-06-15 # 기간 지정, 4개 매장
  python spoton_report.py 2026-06-09 2026-06-15 KCS  # KCS만

출력: ./output/spoton_sales_YYYY-MM-DD[_to_YYYY-MM-DD].xlsx
"""
import sys, io, os, json, datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "https://restaurantreports.spoton.com"
STATE = os.path.join(HERE, "spoton_state.json")
USER = "eg231478@samchully.co.kr"
PW = "Cjk851219!"
WEEKLY_DASH_ID = "54f4dd7b22e45a0000c07bf0"

# (표시명, location_key)
ALL_LOCATIONS = [
    ("Kalbi Social Club - Irvine", "1818994375889661568"),
    ("Kalbi Social Club - Brea",   "1912595012202665856"),
    ("Robata Wasa - Irvine",       "1699495389898777408"),
    ("Robata Wasa - Brea",         "1910722970924225280"),
]
ALIAS = {"KCS": "Kalbi Social Club - Irvine"}

WEEKDAY_KR = ["월", "화", "수", "목", "금", "토", "일"]


def date_range(start, end):
    return {
        "id": 0, "rangeType": "custom", "name": "Custom", "group": "Days",
        "startDate": f"{start}T15:00:00.000Z",
        "endDate":   f"{end}T15:00:00.000Z",
    }


def fetch_guest_sales(page, bearer, loc_key, start, end):
    params = {"date_range": date_range(start, end), "location_key": loc_key}
    url = f"{BASE}/api/reports/{WEEKLY_DASH_ID}?parameters=" + json.dumps(params, separators=(",", ":"))
    res = page.evaluate(
        """async ({url, bearer}) => {
            const r = await fetch(url, {headers: {Authorization: bearer, Accept: 'application/json'}});
            return {status: r.status, body: await r.text()};
        }""",
        {"url": url, "bearer": bearer},
    )
    if res["status"] != 200:
        return []
    data = json.loads(res["body"])
    for blk in data:
        if isinstance(blk, dict) and blk.get("query", {}).get("name") == "Weekly Dashboard - Guest Sales":
            return blk["data"]
    return []


def get_page_with_bearer(ctx_holder, browser):
    ctx = browser.new_context(storage_state=STATE if os.path.exists(STATE) else None)
    ctx_holder["ctx"] = ctx
    page = ctx.new_page()
    bearer_holder = {}
    def on_resp(resp):
        if "/api/reports/" in resp.url:
            a = resp.request.headers.get("authorization", "")
            if a.startswith("Bearer "):
                bearer_holder["b"] = a
    page.on("response", on_resp)
    nav = f"{BASE}/restaurant-reporting/interactive-reports/weekly-dashboard/?location_key={ALL_LOCATIONS[0][1]}"
    page.goto(nav, wait_until="domcontentloaded", timeout=60000)

    # wait until the page settles on either the login form or the app
    need_login = False
    for _ in range(40):
        page.wait_for_timeout(1000)
        try:
            if page.query_selector("input[name='identifier']"):
                need_login = True
                break
            host = page.evaluate("() => location.host + location.pathname")
            if "restaurantreports.spoton.com" in host and "login" not in host:
                break
        except Exception:
            continue

    if need_login:
        page.wait_for_selector("input[name='identifier']", timeout=30000)
        page.fill("input[name='identifier']", USER)
        page.fill("input[name='credentials.passcode']", PW)
        try: page.check("input[name='rememberMe']")
        except Exception: pass
        page.click("input[type='submit']")
        page.wait_for_function(
            "() => location.host.includes('restaurantreports.spoton.com') && !location.pathname.includes('login')",
            timeout=60000)
        page.goto(nav, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(9000)
    ctx.storage_state(path=STATE)
    return page, bearer_holder.get("b")


def fmt_date(dk):
    s = str(dk)
    d = datetime.date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return d, f"{d.year}-{d.month:02d}-{d.day:02d}", WEEKDAY_KR[d.weekday()]


def build_excel(records, out_path, title):
    wb = Workbook()
    ws = wb.active
    ws.title = "매출-객수"
    headers = ["날짜", "요일", "매장", "순매출(USD)", "객수", "거래수", "객단가(USD)"]

    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, size=14)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    hr = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = head_fill; cell.font = head_font; cell.alignment = center; cell.border = border

    r = hr + 1
    for rec in records:
        d, ds, wd = fmt_date(rec["business_date_key"])
        ns = round(rec["net_sales"], 2)
        guests = rec["guest_count"]
        checks = rec["check_count"]
        avg = round(ns / guests, 2) if guests else 0
        vals = [ds, wd, rec["_loc"], ns, guests, checks, avg]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center if c != 3 else Alignment(horizontal="left", vertical="center")
            if c in (4, 7):
                cell.number_format = '#,##0.00'
            if c in (5, 6):
                cell.number_format = '#,##0'
        r += 1

    # totals
    if records:
        ws.cell(row=r, column=1, value="합계").font = Font(bold=True)
        for c in (4, 5, 6):
            col = get_column_letter(c)
            cell = ws.cell(row=r, column=c, value=f"=SUM({col}{hr+1}:{col}{r-1})")
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00' if c == 4 else '#,##0'
            cell.border = border
        ws.cell(row=r, column=3, value="").border = border
        for c in (1, 2, 3, 7):
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=1).font = Font(bold=True)

    widths = [12, 6, 26, 14, 8, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


def main():
    args = sys.argv[1:]
    today = datetime.date.today()
    if len(args) >= 2:
        start, end = args[0], args[1]
    else:
        y = today - datetime.timedelta(days=1)
        start = end = y.isoformat()

    locs = ALL_LOCATIONS
    if len(args) >= 3:
        want = ALIAS.get(args[2], args[2])
        locs = [l for l in ALL_LOCATIONS if l[0] == want]

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        holder = {}
        page, bearer = get_page_with_bearer(holder, browser)
        if not bearer:
            print("ERROR: 로그인/토큰 획득 실패"); browser.close(); sys.exit(1)

        records = []
        for name, key in locs:
            rows = fetch_guest_sales(page, bearer, key, start, end)
            for row in rows:
                row["_loc"] = name
                records.append(row)
            print(f"{name}: {len(rows)} 일치 데이터")
        browser.close()

    # sort by date then location
    records.sort(key=lambda x: (x["business_date_key"], x["_loc"]))

    if start == end:
        out = os.path.join(HERE, "output", f"spoton_sales_{start}.xlsx")
        title = f"SpotOn 일별 매출·객수 — {start}"
    else:
        out = os.path.join(HERE, "output", f"spoton_sales_{start}_to_{end}.xlsx")
        title = f"SpotOn 일별 매출·객수 — {start} ~ {end}"
    build_excel(records, out, title)
    print("SAVED:", out)
    print("ROWS:", len(records))


if __name__ == "__main__":
    main()
